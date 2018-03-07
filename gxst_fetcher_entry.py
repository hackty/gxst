#!/usr/bin/python
# -*- coding: UTF-8 -*-
import time
from openpyxl import Workbook, load_workbook

from selenium import webdriver
from selenium.webdriver import DesiredCapabilities
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

prefix = 'target'
suffix = '.xlsx'
target = prefix+suffix

def get_url(name):


    driver = webdriver.Firefox()
    # driver.implicitly_wait(10)
    driver.set_page_load_timeout(120)
    driver.get('http://www.gsxt.gov.cn')

    locator = (By.ID, 'keyword')
    try:
        WebDriverWait(driver, 60, 0.5).until(EC.presence_of_element_located(locator))
        time.sleep(2)
        input_ele = driver.find_element_by_id('keyword')
        input_ele.clear()
        input_ele.send_keys(name)
    except Exception, e:
        driver.close()
        return False

    locator = (By.ID, 'btn_query')
    try:
        WebDriverWait(driver, 60, 0.5).until(EC.element_to_be_clickable(locator))
        time.sleep(2)
        btn_ele = driver.find_element_by_id('btn_query')
        btn_ele.click()
        time.sleep(3)
    except Exception,e:
        driver.close()
        return False

    locator = (By.CSS_SELECTOR, '.search_list_item.db')
    try:
        WebDriverWait(driver, 60, 0.5).until(EC.presence_of_element_located(locator))
        line_item = driver.find_elements_by_css_selector('.search_list_item.db')
        url = line_item[0].get_attribute('href')
        driver.close()
        return url

    except Exception,e:
        driver.close()
        return False


def main():
    start = 1
    wb = load_workbook(target)
    sheet = wb['Squirrel SQL Export']
    # wb.close()

    for i in range(start,sheet.max_row):
        name = sheet.cell(row=i, column=2).value
        url = sheet.cell(row=i, column=3).value
        if url is not None:
           continue
        else:
            url = get_url(name)
            sheet.cell(row=i, column=3).value=url
            wb.save(target)



if __name__ == '__main__':

    main()
